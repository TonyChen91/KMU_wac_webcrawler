from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
#https://wac.kmu.edu.tw/stu/stuaca/stuqucrs.php
#from bs4 import BeautifulSoup
#import requests
import time
from openpyxl import Workbook, load_workbook
#import json
#可變變數地區路徑
PATH = 'C:/Users/tony/Desktop/學校作業/程式/Day3/chromedriver.exe'
##
driver = webdriver.Chrome(PATH)
driver.get(
    "https://wac.kmu.edu.tw/loginnew.php?usertype=stu&PNO=/stu/stuaca/stuqucrs.php&paras=")
#登陸
time.sleep(0.2)
login = driver.find_element(By.ID, "userid")
login.send_keys("110001047")
login = driver.find_element(By.ID, "password")
login.send_keys("###你的wac密碼###")
login.send_keys(Keys.RETURN)
time.sleep(0.2)

#儲存cookie
#cookie = driver.get_cookies()
#with open('wac.json', 'w') as f:
#    f.write(json.dumps(cookie))
#time.sleep(0.5)
#with open('wac.json', 'r') as f:
#    dd=json.loads(f.read())
#for c in dd:
#    driver.add_cookie(c)
#time.sleep(0.5)
#driver.refresh()

#選單選擇
driver.get("https://wac.kmu.edu.tw/stu/stuaca/stuqucrs.php?FIELDNAME2=&_CD_CONBT=%ACd%B8%DF&_CD_DEPTNO=996&_CD_CHINESECO=&_CD_SGRADE=1&_CD_SM=3&_CD_TEA=&_CD_LANG=&_CD_COLYN=N&d_CurRec=0&d_UPDMode=&d_InSearch=&d_SearchSQL=")
#丟給beautifulsoup
#current_url = driver.current_url
#header={"User-Agent":"Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/98.0.4758.82 Mobile Safari/537.36"}
#response = requests.get(current_url, headers=header)
#print(response)
#soup = BeautifulSoup(response.content, 'html.parser')
#print(soup)

#資料分析
p = 1 #第一頁
excelp = 1
wb = load_workbook("GGG.xlsx")
excelname = str(time.time())
wb.create_sheet(excelname)
ws = wb[excelname]
while p < 9: #高醫總共九頁的選課頁面而已
    k = 1
    while k < 11:
        courses = driver.find_element_by_xpath(
            '//*[@id="d_view_id"]/table[1]/tbody/tr[%s]/td[7]/font' % k)
        codes = driver.find_element_by_xpath(
            '//*[@id="d_view_id"]/table[1]/tbody/tr[%s]/td[3]/font/a' % k)
        chances = driver.find_element_by_xpath(
            '//*[@id="d_view_id"]/table[1]/tbody/tr[%s]/td[12]/b/font' % k)
        categories = driver.find_element_by_xpath(
            '//*[@id="d_view_id"]/table[1]/tbody/tr[%s]/td[9]/font' % k)
        #print(courses.text)
        #print(chances.text)
        ws["A"+str(excelp)].value = courses.text
        ws["B"+str(excelp)].value = chances.text
        ws["C"+str(excelp)].value = codes.text
        ws["D"+str(excelp)].value = categories.text
        #print(ws["A"+str(excelp)].value)
        k += 1
        excelp += 1
    #換頁
    nextpage = driver.find_element_by_id("d_id_nextp_1")
    nextpage.click()
    time.sleep(0.2)
    p += 1
#time.sleep(3)

#分析資料
#search = driver.find_element_by_id("d_id_nextp_1")
#search.click()
#time.sleep(3)

#寫入EXCEL檔案
wb.save('GGG.xlsx')
driver.quit()
